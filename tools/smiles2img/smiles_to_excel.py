#!/usr/bin/env python3
"""CSVのSMILES列を構造式PNGに変換し、Excel(.xlsx)の同じ列に埋め込む。

使い方:
    python smiles_to_excel.py input.csv output.xlsx --column E
    python smiles_to_excel.py input.csv output.xlsx --column SMILES
    python smiles_to_excel.py input.tsv output.xlsx --column E --delimiter tab
    # SMILESはE列のまま残し、画像はF列に貼る
    python smiles_to_excel.py input.csv output.xlsx --column E --image-column F
    # 1ファイルあたり3000行ずつに分割して出力(既定は5000行)
    python smiles_to_excel.py input.csv output.xlsx --column E --chunk-size 3000

件数が多い場合は必ず --chunk-size に応じて output_part01.xlsx, output_part02.xlsx ...
のように分割保存される(openpyxlは画像入りワークブックを2回save()すると壊れるため、
1ワークブックにつき保存は1回のみに限定している)。
"""
import argparse
import csv
import io
import sys
import traceback
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import column_index_from_string, get_column_letter
from rdkit import Chem
from rdkit.Chem import Draw

IMG_SIZE = (200, 200)
CELL_ROW_HEIGHT = 150  # points
CELL_COL_WIDTH = 28  # excel column width units

# 日本語Excelで保存されたCSVはShift-JIS(CP932)になることが多いため、
# UTF-8で読めなければ順に試す
CANDIDATE_ENCODINGS = ["utf-8-sig", "cp932", "euc_jp"]
CANDIDATE_DELIMITERS = [",", "\t", ";", "|"]
DELIMITER_ALIASES = {"tab": "\t", "comma": ",", "semicolon": ";"}


def detect_delimiter(sample_text):
    try:
        dialect = csv.Sniffer().sniff(sample_text, delimiters="".join(CANDIDATE_DELIMITERS))
        return dialect.delimiter
    except csv.Error:
        first_line = sample_text.splitlines()[0] if sample_text else ""
        counts = {d: first_line.count(d) for d in CANDIDATE_DELIMITERS}
        best = max(counts, key=counts.get)
        return best if counts[best] > 0 else ","


def read_csv_rows(path, encoding=None, delimiter=None):
    encodings = [encoding] if encoding else CANDIDATE_ENCODINGS
    for enc in encodings:
        try:
            text = path.read_text(encoding=enc)
        except UnicodeDecodeError:
            continue
        delim = delimiter or detect_delimiter(text)
        rows = list(csv.reader(io.StringIO(text), delimiter=delim))
        delim_label = "TAB" if delim == "\t" else repr(delim)
        print(f"[info] 文字コード '{enc}' / 区切り文字 {delim_label} として読み込みました", file=sys.stderr)
        return rows
    raise SystemExit(
        f"CSVの文字コードを判定できませんでした(試した候補: {encodings})。"
        f" --encoding で明示的に指定してください。"
    )


def resolve_column_index(header, column_arg):
    """--column に列名(例: SMILES)か列記号(例: E)のどちらが来ても列インデックス(0始まり)を返す"""
    if column_arg in header:
        return header.index(column_arg)
    try:
        return column_index_from_string(column_arg.upper()) - 1
    except ValueError:
        pass
    raise SystemExit(f"列 '{column_arg}' が見つかりません。ヘッダー: {header}")


def smiles_to_png_bytes(smiles, colors=16):
    mol = Chem.MolFromSmiles(smiles)
    if mol is None:
        raise ValueError("SMILES解析失敗(RDKitがNoneを返した)")
    pil_img = Draw.MolToImage(mol, size=IMG_SIZE)
    if colors:
        # 構造式はほぼ白黒少数色なので、パレット化するだけでPNGが大幅に軽くなる
        pil_img = pil_img.convert("P", palette=1, colors=colors)
    buf = io.BytesIO()
    pil_img.save(buf, format="PNG", optimize=True)
    buf.seek(0)
    return buf


def chunk_output_path(output_path, part_no, total_parts):
    if total_parts == 1:
        return output_path
    width = len(str(total_parts))
    return output_path.with_name(f"{output_path.stem}_part{part_no:0{width}d}{output_path.suffix}")


def write_chunk(chunk_rows, header, smiles_idx, image_idx, image_col_letter,
                 overwrite_smiles_text, clear_text, image_colors, out_path, row_offset, total,
                 progress_every, failed):
    wb = Workbook()
    ws = wb.active

    for col_i, val in enumerate(header):
        ws.cell(row=1, column=col_i + 1, value=val)
    if image_idx >= len(header) or not header[image_idx].strip():
        ws.cell(row=1, column=image_idx + 1, value="Structure")
    ws.column_dimensions[image_col_letter].width = CELL_COL_WIDTH

    for local_i, row in enumerate(chunk_rows):
        global_i = row_offset + local_i + 1  # 1始まりの通し番号(データ行のみ)
        row_i = local_i + 2  # このチャンク内でのExcel行番号
        for col_i, val in enumerate(row):
            ws.cell(row=row_i, column=col_i + 1, value=val)
        ws.row_dimensions[row_i].height = CELL_ROW_HEIGHT

        smiles = row[smiles_idx].strip() if smiles_idx < len(row) else ""
        if smiles:
            try:
                png_buf = smiles_to_png_bytes(smiles, colors=image_colors)
                cell_ref = f"{image_col_letter}{row_i}"
                if overwrite_smiles_text and clear_text:
                    ws[cell_ref] = None  # SMILESテキストを消して画像に差し替え
                img = XLImage(png_buf)
                img.width, img.height = IMG_SIZE
                img.anchor = cell_ref
                ws.add_image(img)
            except Exception as exc:
                # 1件の失敗でチャンク全体を止めない
                failed.append((global_i, smiles, f"{type(exc).__name__}: {exc}"))
                traceback.print_exc(file=sys.stderr)

        if progress_every and global_i % progress_every == 0:
            print(f"[progress] {global_i}/{total}行処理(失敗{len(failed)}件)", file=sys.stderr)

    wb.save(out_path)
    print(f"[part] {out_path} に保存しました({len(chunk_rows)}行)", file=sys.stderr)


def main():
    parser = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument("input_csv", type=Path)
    parser.add_argument("output_xlsx", type=Path)
    parser.add_argument(
        "--column",
        default="E",
        help="SMILESが入っている列。列記号(E)またはヘッダー名(SMILES)。既定: E",
    )
    parser.add_argument(
        "--image-column",
        default=None,
        help="画像を貼る列。列記号(F)またはヘッダー名。省略時は--columnと同じ列",
    )
    parser.add_argument(
        "--clear-text",
        action="store_true",
        help="画像をSMILES列自体に貼る場合(--image-column省略時)、"
        "元のSMILESテキストを消してから画像を貼る。既定ではテキストは残したまま"
        "画像を上に重ねる(セルの値としては残るが、画像の下に隠れて見えなくなる)",
    )
    parser.add_argument(
        "--encoding",
        default=None,
        help="CSVの文字コードを明示指定(例: cp932)。省略時は自動判定",
    )
    parser.add_argument(
        "--delimiter",
        default=None,
        help="区切り文字を明示指定(例: tab, ';' )。省略時は自動判定",
    )
    parser.add_argument(
        "--chunk-size",
        type=int,
        default=5000,
        help="1つのxlsxに詰める最大行数。超える場合は_partNNに分割保存する。"
        "0を指定すると分割せず1ファイルにまとめる(件数が多いと非推奨)。既定: 5000",
    )
    parser.add_argument(
        "--progress-every",
        type=int,
        default=500,
        help="この行数ごとに進捗をstderrへ出す。既定: 500行ごと",
    )
    parser.add_argument(
        "--image-colors",
        type=int,
        default=16,
        help="画像PNGの色数をこの数に減色してファイルサイズを圧縮する(構造式は少数色で足りる)。"
        "0を指定すると減色せずフルカラーで保存する。既定: 16色",
    )
    args = parser.parse_args()

    delimiter = DELIMITER_ALIASES.get(args.delimiter, args.delimiter) if args.delimiter else None
    rows = read_csv_rows(args.input_csv, encoding=args.encoding, delimiter=delimiter)
    if not rows:
        raise SystemExit("CSVが空です")

    header, data_rows = rows[0], rows[1:]
    smiles_idx = resolve_column_index(header, args.column)
    image_idx = resolve_column_index(header, args.image_column) if args.image_column else smiles_idx
    overwrite_smiles_text = image_idx == smiles_idx
    image_col_letter = get_column_letter(image_idx + 1)

    total = len(data_rows)
    chunk_size = args.chunk_size if args.chunk_size and args.chunk_size > 0 else total
    chunks = [data_rows[i : i + chunk_size] for i in range(0, total, chunk_size)] or [[]]
    total_parts = len(chunks)

    failed = []
    output_files = []
    for part_no, chunk_rows in enumerate(chunks, start=1):
        out_path = chunk_output_path(args.output_xlsx, part_no, total_parts)
        output_files.append(out_path)
        row_offset = (part_no - 1) * chunk_size
        write_chunk(
            chunk_rows, header, smiles_idx, image_idx, image_col_letter,
            overwrite_smiles_text, args.clear_text, args.image_colors, out_path, row_offset, total,
            args.progress_every, failed,
        )

    print(f"書き出し完了: {total}行 / {total_parts}ファイル, 失敗{len(failed)}件")
    for out_path in output_files:
        print(f"  -> {out_path}")
    for row_i, smiles, reason in failed:
        print(f"  [warn] row {row_i}: {reason} -> {smiles!r}", file=sys.stderr)


if __name__ == "__main__":
    main()
